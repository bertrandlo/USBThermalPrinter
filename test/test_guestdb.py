import sqlite3
import turtle
import unittest
from collections import deque

from jpype._core import startJVM
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


def load_stroke_count():
    report = {}
    conn = sqlite3.connect('utf8stroke.db')
    c = conn.cursor()
    c.execute("SELECT strokenum, word FROM stroke")
    for row in c.fetchall():
        report[row[1]] = row[0]
    return report


def as_text(value):
    if value is None:
        return ""
    return str(value)


class GuestInfo:
    stroke_db = load_stroke_count()

    def __init__(self, row: turtle):
        self.family_name = row[1][0:1]
        self.name = row[1]
        self.sorting_weight = row[5]
        self.family_name_stroke_count = 0
        self.phone = row[2]
        self.table_id = row[3]
        self.note = row[4]
        if self.family_name in self.stroke_db.keys():
            self.family_name_stroke_count = self.stroke_db.get(self.family_name)


def query_database():
    report = {}
    conn = sqlite3.connect('guests.db')
    c = conn.cursor()
    c.execute("SELECT * FROM guests ORDER BY sortingweight")
    for row in c:
        guest_info = GuestInfo(row)
        if guest_info.sorting_weight == 0:
            continue
        # print("{}\t\t\t{}".format(guest_info.family_name, guest_info.sorting_weight))
        if guest_info.family_name not in report.keys():
            report[guest_info.family_name] = deque()

        deque_guest = report.get(guest_info.family_name)
        deque_guest.append(guest_info)
    conn.close()

    return report


class TestGuestDB(unittest.TestCase):

    def test_open_sqlite_file(self):
        startJVM("C:/Users/brt/.jdks/adopt-openjdk-11.0.11/bin/server/jvm.dll", "-ea")

        import asposecells.api

        report = query_database()
        # 依據姓氏排序
        family_name_stroke_count_db = load_stroke_count()
        family_name_list = sorted(report.keys(), key=lambda fn: family_name_stroke_count_db.get(fn))

        wb = Workbook()
        ws = wb.active

        darklight = NamedStyle(name="darklight")
        darklight.font = Font(size=12)
        darklight.fill = PatternFill("solid", fgColor="DDDDDD")

        highlight = NamedStyle(name="highlight")
        highlight.font = Font(size=12)
        colors = [darklight, highlight]

        for i, family_name in enumerate(family_name_list):
            # print("{} {}".format(family_name, len(report.get(family_name))))
            members = sorted(list(report.get(family_name)), key=lambda m: m.sorting_weight)

            guest: GuestInfo
            for guest in members:
                # print("{}\t\t{}\t\t{}".format(guest.family_name, guest.name, guest.sorting_weight))
                ws.append([guest.name, guest.phone, guest.table_id, guest.note])
                for col_index in range(1, 5):
                    ws.cell(row=ws.max_row, column=col_index).style = colors[i % 2]
            ws.append([" "])

        width = [0, 12, 24, 12, 50]   # 0 for padding, no meaning

        for i, column_cells in enumerate(ws.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width[i]

        ws.page_margins.left = 1.5 / 2.54
        ws.page_margins.right = 0.5 / 2.54
        ws.page_margins.top = 0.5 / 2.54
        ws.page_margins.bottom = 0.5 / 2.54

        wb.save("sample.xlsx")

        workbook = asposecells.api.Workbook("sample.xlsx")
        workbook.save("sample.pdf", asposecells.api.SaveFormat.PDF)


class TestReportLab(unittest.TestCase):

    pdfmetrics.registerFont(TTFont('msjh', 'msjh.ttc'))  # 指定字型

    def test_print(self):
        c = canvas.Canvas("123.pdf")
        c.setFont('msjh', 20)
        message = "Hello,你好，歡迎學習reportlab製作pdf！"
        # 注意下面坐標x和y，當（0,0）是代表左下角 #一張pdf的高800，所以第一行800，再大就出
        c.drawString(50, 800, message.encode('utf-8'))
        c.setFillColorRGB(1, 0, 0)
        c.rect(5, 5, 652, 792, fill=1)
        c.showPage()
        c.save()

    def test_table(self):
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        elements = []
        data = []

        report = query_database()
        # 依據姓氏排序
        family_name_stroke_count_db = load_stroke_count()
        family_name_list = sorted(report.keys(), key=lambda fn: family_name_stroke_count_db.get(fn))
        blank_rows_list = []
        row_index = 0

        for i, family_name in enumerate(family_name_list):
            # print("{} {}".format(family_name, len(report.get(family_name))))
            members = sorted(list(report.get(family_name)), key=lambda m: m.sorting_weight)

            guest: GuestInfo
            for guest in members:
                # print("{}\t\t{}\t\t{}".format(guest.family_name, guest.name, guest.sorting_weight))
                data.append([guest.name, guest.phone, guest.table_id, guest.note])
                row_index = row_index + 1

            blank_rows_list.append(row_index)
            data.append(["", "", "", ""])
            row_index = row_index + 1

        table = Table(data, [2.0*cm, 4.5*cm, 1.5*cm, 11.5*cm], len(data)*[0.6*cm])
        base_style = [
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.green),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
            ('FONT', (0, 0), (-1, -1), 'msjh')
        ]
        for i in blank_rows_list:
            base_style.append(('BACKGROUND', (0, i), (-1, i), colors.lightblue))

        table.setStyle(TableStyle(base_style))

        elements.append(table)
        doc = SimpleDocTemplate("123.pdf", rightMargin=1.0*cm, leftMargin=1.0*cm, topMargin=1.0*cm, bottomMargin=1.0*cm)
        doc.build(elements)
